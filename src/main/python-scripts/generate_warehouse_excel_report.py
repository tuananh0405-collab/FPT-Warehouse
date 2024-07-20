# import pandas as pd
# import json
# import sys
# from datetime import datetime
#
# def generate_warehouse_excel_report(input_json_file, output_excel_file):
#     with open(input_json_file, 'r') as f:
#         snapshots = json.load(f)
#
#     if not snapshots:
#         print("No data found in JSON file.")
#         return
#
#     data = {}
#     for snapshot in snapshots:
#         warehouse = snapshot['warehouse']
#         product = snapshot['product']
#         zone = snapshot['zone']
#         quantity = snapshot['quantity']
#         snapshot_date = datetime.strptime(snapshot['snapshotDate'], '%Y-%m-%d').strftime('%Y-%m-%d')
#
#         if (warehouse, product, zone) not in data:
#             data[(warehouse, product, zone)] = []
#         data[(warehouse, product, zone)].append((quantity, snapshot_date))
#
#     formatted_data = []
#     for (warehouse, product, zone), records in data.items():
#         row = {
#             'Warehouse': warehouse,
#             'Product Name': product,
#             'Zone Name': zone,
#         }
#         for i, (quantity, snapshot_date) in enumerate(records):
#             row[f'Quantity {i+1}'] = quantity
#             row[f'Snapshot Date {i+1}'] = snapshot_date
#         formatted_data.append(row)
#
#     df = pd.DataFrame(formatted_data)
#     df.to_excel(output_excel_file, index=False)
#     print(f"Excel report generated at {output_excel_file}")
#
# if __name__ == "__main__":
#     if len(sys.argv) != 3:
#         print("Usage: python generate_warehouse_excel_report.py <input_json_file> <output_excel_file>")
#         sys.exit(1)
#
#     input_json_file = sys.argv[1]
#     output_excel_file = sys.argv[2]
#     generate_warehouse_excel_report(input_json_file, output_excel_file)
import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def generate_warehouse_excel_report(input_json_file, output_excel_file):
    with open(input_json_file, 'r') as f:
        snapshots = json.load(f)

    if not snapshots:
        print("No data found in JSON file.")
        return

    data = {}
    for snapshot in snapshots:
        warehouse = snapshot['warehouse']
        product = snapshot['product']
        zone = snapshot['zone']
        quantity = snapshot['quantity']
        snapshot_date = datetime.strptime(snapshot['snapshotDate'], '%Y-%m-%d').strftime('%Y-%m-%d')

        if (warehouse, product, zone) not in data:
            data[(warehouse, product, zone)] = []
        data[(warehouse, product, zone)].append((quantity, snapshot_date))

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Warehouse Report"

    # Write the header
    headers = ['Warehouse', 'Product Name', 'Zone Name']
    max_records = max(len(records) for records in data.values())
    for i in range(max_records):
        headers.extend([f'Quantity {i+1}', f'Snapshot Date {i+1}'])
    sheet.append(headers)

    # Write the data
    for (warehouse, product, zone), records in data.items():
        row = [warehouse, product, zone]
        for quantity, snapshot_date in records:
            row.extend([quantity, snapshot_date])
        sheet.append(row)

    # Adjust column widths
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_size = True

    workbook.save(output_excel_file)
    print(f"Excel report generated at {output_excel_file}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_warehouse_excel_report.py <input_json_file> <output_excel_file>")
        sys.exit(1)

    input_json_file = sys.argv[1]
    output_excel_file = sys.argv[2]
    generate_warehouse_excel_report(input_json_file, output_excel_file)
